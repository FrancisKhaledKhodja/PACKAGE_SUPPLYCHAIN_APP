import os
import tempfile
from datetime import datetime

import polars as pl
from flask import jsonify, send_file, request
from openpyxl import Workbook

from . import bp
from supplychain_app.constants import path_r, folder_gestion_pr, path_lmline, path_datan, folder_name_app


CARNET_CHRONOPOST_DIR = r"R:\24-DPR\11-Applications\04-Gestion_Des_Points_Relais\Data\GESTION_PR\CARNET_CHRONOPOST"
CHRONO_FUSION_DIR = r"R:\24-DPR\11-Applications\04-Gestion_Des_Points_Relais\Data\CHRONOPOST\2_C9_C13_EXCEL_FUSION"


ARTICLE_REQUESTS_DEMANDES_DIR = r"\\apps\Vol1\Data\011-BO_XI_entrees\07-DOR_DP\Sorties\FICHIERS_REFERENTIEL_ARTICLE\DEMANDES"


def _latest_file_in_dir(directory: str, pattern_ext: tuple = (".xlsx", ".xls", ".csv")):
    """Return the lexicographically largest file in directory matching given extensions."""
    try:
        names = [
            f for f in os.listdir(directory)
            if f.lower().endswith(pattern_ext)
        ]
        if not names:
            return None
        best = max(names, key=lambda s: s.lower())
        return os.path.join(directory, best)
    except Exception:
        return None


@bp.get("/")
def list_latest_files():
    """Return the latest available business files for download (no parquets)."""
    # Annuaire PR
    annuaire_dir = os.path.join(path_r, folder_gestion_pr, "ANNUAIRE_PR")
    latest_annuaire = _latest_file_in_dir(annuaire_dir)

    # LM2S
    latest_lm2s = _latest_file_in_dir(path_lmline)

    # Chronopost fusionné (nouveau répertoire dédié)
    latest_chrono = _latest_file_in_dir(CHRONO_FUSION_DIR)

    # Carnet Chronopost
    latest_carnet = _latest_file_in_dir(CARNET_CHRONOPOST_DIR)

    # Stock final parquet (pour export CSV)
    stock_final_parquet = os.path.join(path_datan, folder_name_app, "stock_final.parquet")
    latest_stock_final = stock_final_parquet if os.path.exists(stock_final_parquet) else None

    def _info(path: str | None):
        if not path:
            return {"available": False, "filename": None}
        return {"available": True, "filename": os.path.basename(path)}

    return jsonify({
        "annuaire_pr": _info(latest_annuaire),
        "chronopost_fusionne": _info(latest_chrono),
        "lm2s": _info(latest_lm2s),
        "carnet_chronopost": _info(latest_carnet),
        "stock_final_csv": _info(latest_stock_final),
    })


@bp.get("/annuaire_pr")
def download_annuaire_pr_api():
    annuaire_dir = os.path.join(path_r, folder_gestion_pr, "ANNUAIRE_PR")
    latest = _latest_file_in_dir(annuaire_dir)
    if not latest:
        return jsonify({"error": "not_found"}), 404
    return send_file(latest, as_attachment=True, download_name=os.path.basename(latest))


@bp.get("/stock_final_csv")
def download_stock_final_csv_api():
    stock_parquet = os.path.join(path_datan, folder_name_app, "stock_final.parquet")
    if not os.path.exists(stock_parquet):
        return jsonify({"error": "not_found"}), 404

    try:
        df = pl.read_parquet(stock_parquet)
    except Exception:
        return jsonify({"error": "read_failed"}), 500

    tmp = tempfile.NamedTemporaryFile(suffix=".csv", delete=False)
    tmp_path = tmp.name
    tmp.close()

    try:
        df.write_csv(tmp_path, separator=";")
    except Exception:
        return jsonify({"error": "export_failed"}), 500

    return send_file(
        tmp_path,
        as_attachment=True,
        download_name="stock_final.csv",
        mimetype="text/csv; charset=utf-8",
    )


@bp.get("/lm2s")
def download_lm2s_api():
    latest = _latest_file_in_dir(path_lmline)
    if not latest:
        return jsonify({"error": "not_found"}), 404
    return send_file(latest, as_attachment=True, download_name=os.path.basename(latest))


@bp.get("/chronopost")
def download_chronopost_api():
    latest = _latest_file_in_dir(CHRONO_FUSION_DIR)
    if not latest:
        return jsonify({"error": "not_found"}), 404
    return send_file(latest, as_attachment=True, download_name=os.path.basename(latest))


@bp.get("/carnet_chronopost")
def download_carnet_chronopost_api():
    latest = _latest_file_in_dir(CARNET_CHRONOPOST_DIR)
    if not latest:
        return jsonify({"error": "not_found"}), 404
    return send_file(latest, as_attachment=True, download_name=os.path.basename(latest))


@bp.post("/demandes/modif_criticite_xlsx")
def save_demande_modif_criticite_xlsx():
    body = request.get_json(silent=True) or {}

    date_demande = (body.get("date_demande") or "").strip()
    demandeur = (body.get("demandeur") or "").strip()
    type_demande = (body.get("type_demande") or "").strip() or "Modification d'une criticité"
    rows = body.get("rows") or []

    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "rows_required"}), 400

    ts = datetime.now().strftime("%Y%m%dT%H%M%S")
    filename = f"dde_modif_art_criticite_{ts}.xlsx"

    try:
        os.makedirs(ARTICLE_REQUESTS_DEMANDES_DIR, exist_ok=True)
    except Exception:
        return jsonify({"error": "mkdir_failed", "dir": ARTICLE_REQUESTS_DEMANDES_DIR}), 500

    out_path = os.path.join(ARTICLE_REQUESTS_DEMANDES_DIR, filename)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "demandes"

        headers = [
            "date_demande",
            "demandeur",
            "type_demande",
            "code_article",
            "feuille_du_catalogue",
            "type_article",
            "libelle_article",
            "criticite_article",
            "nouvelle_criticite_article",
            "causes",
        ]
        ws.append(headers)

        def _s(v):
            return "" if v is None else str(v)

        for r in rows:
            if not isinstance(r, dict):
                continue
            ws.append([
                date_demande,
                demandeur,
                type_demande,
                _s(r.get("code_article")),
                _s(r.get("feuille_du_catalogue")),
                _s(r.get("type_article")),
                _s(r.get("libelle_article")),
                _s(r.get("criticite_article")),
                _s(r.get("nouvelle_criticite_article")),
                _s(r.get("causes")),
            ])

        wb.save(out_path)
    except Exception:
        return jsonify({"error": "write_failed"}), 500

    return jsonify({
        "ok": True,
        "filename": filename,
        "path": out_path,
        "dir": ARTICLE_REQUESTS_DEMANDES_DIR,
    })


@bp.post("/demandes/creation_articles_xlsx")
def save_demande_creation_articles_xlsx():
    body = request.get_json(silent=True) or {}

    date_demande = (body.get("date_demande") or "").strip()
    demandeur = (body.get("demandeur") or "").strip()
    type_demande = (body.get("type_demande") or "").strip() or "Création d'article(s)"
    rows = body.get("rows") or []

    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "rows_required"}), 400

    ts = datetime.now().strftime("%Y%m%dT%H%M%S")
    filename = f"dde_creation_article_{ts}.xlsx"

    try:
        os.makedirs(ARTICLE_REQUESTS_DEMANDES_DIR, exist_ok=True)
    except Exception:
        return jsonify({"error": "mkdir_failed", "dir": ARTICLE_REQUESTS_DEMANDES_DIR}), 500

    out_path = os.path.join(ARTICLE_REQUESTS_DEMANDES_DIR, filename)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "demandes"

        headers = [
            "date_demande",
            "demandeur",
            "type_demande",
            "article_tdf_clt",
            "feuille_du_catalogue",
            "libelle_court",
            "libelle_long",
            "type_article",
            "cycle_vie_production",
            "statut_article",
            "cycle_vie_achat",
            "commentaire_technique",
            "poids_kg",
            "long_m",
            "larg_m",
            "haut_m",
            "article_datacenter",
            "article_hors_norme",
            "peremption",
            "matiere_dangereuse",
            "categorie_md",
            "code_onu",
            "catalogue_consommables",
            "categorie_consommables",
            "sous_categorie_consommables",
            "retour_production",
            "prix_previsionnel",
            "criticite",
            "mnemonique",
            "fabricant",
            "reference_fabricant",
        ]
        ws.append(headers)

        def _s(v):
            return "" if v is None else str(v)

        for r in rows:
            if not isinstance(r, dict):
                continue
            ws.append([
                date_demande,
                demandeur,
                type_demande,
                _s(r.get("article_tdf_clt")),
                _s(r.get("feuille_du_catalogue")),
                _s(r.get("libelle_court")),
                _s(r.get("libelle_long")),
                _s(r.get("type_article")),
                _s(r.get("cycle_vie_production")),
                _s(r.get("statut_article")),
                _s(r.get("cycle_vie_achat")),
                _s(r.get("commentaire_technique")),
                _s(r.get("poids_kg")),
                _s(r.get("long_m")),
                _s(r.get("larg_m")),
                _s(r.get("haut_m")),
                _s(r.get("article_datacenter")),
                _s(r.get("article_hors_norme")),
                _s(r.get("peremption")),
                _s(r.get("matiere_dangereuse")),
                _s(r.get("categorie_md")),
                _s(r.get("code_onu")),
                _s(r.get("catalogue_consommables")),
                _s(r.get("categorie_consommables")),
                _s(r.get("sous_categorie_consommables")),
                _s(r.get("retour_production")),
                _s(r.get("prix_previsionnel")),
                _s(r.get("criticite")),
                _s(r.get("mnemonique")),
                _s(r.get("fabricant")),
                _s(r.get("reference_fabricant")),
            ])

        wb.save(out_path)
    except Exception:
        return jsonify({"error": "write_failed"}), 500

    return jsonify({
        "ok": True,
        "filename": filename,
        "path": out_path,
        "dir": ARTICLE_REQUESTS_DEMANDES_DIR,
    })


@bp.post("/demandes/passage_rebut_xlsx")
def save_demande_passage_rebut_xlsx():
    body = request.get_json(silent=True) or {}

    date_demande = (body.get("date_demande") or "").strip()
    demandeur = (body.get("demandeur") or "").strip()
    type_demande = (body.get("type_demande") or "").strip() or "Demande de passage en REBUT"
    rows = body.get("rows") or []

    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "rows_required"}), 400

    ts = datetime.now().strftime("%Y%m%dT%H%M%S")
    filename = f"dde_passage_rebut_{ts}.xlsx"

    try:
        os.makedirs(ARTICLE_REQUESTS_DEMANDES_DIR, exist_ok=True)
    except Exception:
        return jsonify({"error": "mkdir_failed", "dir": ARTICLE_REQUESTS_DEMANDES_DIR}), 500

    out_path = os.path.join(ARTICLE_REQUESTS_DEMANDES_DIR, filename)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "demandes"

        headers = [
            "date_demande",
            "demandeur",
            "type_demande",
            "code_article",
            "libelle_article",
            "feuille_du_catalogue",
            "type_article",
            "statut_abrege_article",
            "criticite",
            "cycle_de_vie_de_production_pim",
            "cycle_de_vie_achat",
            "lieu_de_reparation_pim",
            "rma",
            "retour_production",
        ]
        ws.append(headers)

        def _s(v):
            return "" if v is None else str(v)

        for r in rows:
            if not isinstance(r, dict):
                continue
            ws.append([
                date_demande,
                demandeur,
                type_demande,
                _s(r.get("code_article")),
                _s(r.get("libelle_article")),
                _s(r.get("feuille_du_catalogue")),
                _s(r.get("type_article")),
                _s(r.get("statut_abrege_article")),
                _s(r.get("criticite")),
                _s(r.get("cycle_de_vie_de_production_pim")),
                _s(r.get("cycle_de_vie_achat")),
                _s(r.get("lieu_de_reparation_pim")),
                _s(r.get("rma")),
                _s(r.get("retour_production")),
            ])

        wb.save(out_path)
    except Exception:
        return jsonify({"error": "write_failed"}), 500

    return jsonify({
        "ok": True,
        "filename": filename,
        "path": out_path,
        "dir": ARTICLE_REQUESTS_DEMANDES_DIR,
    })


@bp.post("/demandes/modif_achetable_xlsx")
def save_demande_modif_achetable_xlsx():
    body = request.get_json(silent=True) or {}

    date_demande = (body.get("date_demande") or "").strip()
    demandeur = (body.get("demandeur") or "").strip()
    type_demande = (body.get("type_demande") or "").strip() or "Modification du statut achetable / non achetable"
    rows = body.get("rows") or []

    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "rows_required"}), 400

    ts = datetime.now().strftime("%Y%m%dT%H%M%S")
    filename = f"dde_modif_art_achetable_{ts}.xlsx"

    try:
        os.makedirs(ARTICLE_REQUESTS_DEMANDES_DIR, exist_ok=True)
    except Exception:
        return jsonify({"error": "mkdir_failed", "dir": ARTICLE_REQUESTS_DEMANDES_DIR}), 500

    out_path = os.path.join(ARTICLE_REQUESTS_DEMANDES_DIR, filename)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "demandes"

        headers = [
            "date_demande",
            "demandeur",
            "type_demande",
            "code_article",
            "feuille_du_catalogue",
            "type_article",
            "libelle_article",
            "statut_abrege_article",
            "nouveau_statut_abrege_article",
            "causes",
        ]
        ws.append(headers)

        def _s(v):
            return "" if v is None else str(v)

        for r in rows:
            if not isinstance(r, dict):
                continue
            ws.append([
                date_demande,
                demandeur,
                type_demande,
                _s(r.get("code_article")),
                _s(r.get("feuille_du_catalogue")),
                _s(r.get("type_article")),
                _s(r.get("libelle_article")),
                _s(r.get("statut_abrege_article")),
                _s(r.get("nouveau_statut_abrege_article")),
                _s(r.get("causes")),
            ])

        wb.save(out_path)
    except Exception:
        return jsonify({"error": "write_failed"}), 500

    return jsonify({
        "ok": True,
        "filename": filename,
        "path": out_path,
        "dir": ARTICLE_REQUESTS_DEMANDES_DIR,
    })


@bp.post("/demandes/equivalence_xlsx")
def save_demande_equivalence_xlsx():
    body = request.get_json(silent=True) or {}

    date_demande = (body.get("date_demande") or "").strip()
    demandeur = (body.get("demandeur") or "").strip()
    type_demande = (body.get("type_demande") or "").strip() or "Déclaration d'une équivalence"
    rows = body.get("rows") or []

    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "rows_required"}), 400

    ts = datetime.now().strftime("%Y%m%dT%H%M%S")
    filename = f"dde_equivalence_{ts}.xlsx"

    try:
        os.makedirs(ARTICLE_REQUESTS_DEMANDES_DIR, exist_ok=True)
    except Exception:
        return jsonify({"error": "mkdir_failed", "dir": ARTICLE_REQUESTS_DEMANDES_DIR}), 500

    out_path = os.path.join(ARTICLE_REQUESTS_DEMANDES_DIR, filename)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "demandes"

        headers = [
            "date_demande",
            "demandeur",
            "type_demande",
            "code_article",
            "libelle_article",
            "code_equivalent",
            "libelle_article_equivalent",
            "type_equivalence",
            "reciprocite",
            "rang",
        ]
        ws.append(headers)

        def _s(v):
            return "" if v is None else str(v)

        for r in rows:
            if not isinstance(r, dict):
                continue
            ws.append([
                date_demande,
                demandeur,
                type_demande,
                _s(r.get("code_article")),
                _s(r.get("libelle_article")),
                _s(r.get("code_equivalent")),
                _s(r.get("libelle_article_equivalent")),
                _s(r.get("type_equivalence")),
                _s(r.get("reciprocite")),
                _s(r.get("rang")),
            ])

        wb.save(out_path)
    except Exception:
        return jsonify({"error": "write_failed"}), 500

    return jsonify({
        "ok": True,
        "filename": filename,
        "path": out_path,
        "dir": ARTICLE_REQUESTS_DEMANDES_DIR,
    })

